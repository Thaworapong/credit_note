#V.1.0.0
#First release of the Credit Note application
import os
import json
from datetime import datetime
from PySide6.QtWidgets import (
    QApplication, QWidget, QLabel, QLineEdit, QPushButton, QVBoxLayout,
    QHBoxLayout, QFormLayout, QTableWidget, QTableWidgetItem,
    QDateEdit, QTextEdit, QHeaderView, QMessageBox
)
from PySide6.QtCore import QDate, Qt
from openpyxl import load_workbook
from pythainlp.util import bahttext
from PySide6.QtGui import QIcon

# Mapping Thai digits to Arabic digits
thai_digits = "๐๑๒๓๔๕๖๗๘๙"
arabic_digits = "0123456789"
trans_table = str.maketrans(thai_digits, arabic_digits)

def to_arabic_digits(text: str) -> str:
    return text.translate(trans_table)

class CreditNoteWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Credit Note")
        self.setMinimumSize(800, 600)
        self.setWindowIcon(QIcon(os.path.join(os.path.dirname(__file__), "credit_note.png")))

        self.config_dir = "config"
        self.template_dir = "template"
        self.output_dir = "output"
        self.log_path = os.path.join(self.config_dir, "credit_note_log.json")

        os.makedirs(self.config_dir, exist_ok=True)
        os.makedirs(self.template_dir, exist_ok=True)
        os.makedirs(self.output_dir, exist_ok=True)

        self.load_credit_note_log()
        self.init_ui()

    def load_credit_note_log(self):
        if os.path.exists(self.log_path):
            with open(self.log_path, "r", encoding="utf-8") as f:
                self.credit_note_log = json.load(f)
        else:
            self.credit_note_log = {}

    def save_credit_note_log(self, entry):
        date_key = entry["date"]
        if date_key not in self.credit_note_log:
            self.credit_note_log[date_key] = []
        self.credit_note_log[date_key].append(entry)

        with open(self.log_path, "w", encoding="utf-8") as f:
            json.dump(self.credit_note_log, f, ensure_ascii=False, indent=2)

    def generate_credit_note_no(self):
        now = datetime.now()
        thai_year = now.year + 543
        year_month = f"{thai_year}{now.strftime('%m')}"
        date_key = now.strftime("%Y-%m-%d")

        existing = self.credit_note_log.get(date_key, [])
        next_seq = len(existing) + 1
        num_str = f"{next_seq:03d}"
        return f"CNT{year_month}-{num_str}"

    def init_ui(self):
        layout = QVBoxLayout()
        form_layout = QFormLayout()

        self.credit_note_no = QLineEdit()
        self.credit_note_no.setReadOnly(True)
        self.credit_note_no.setText(self.generate_credit_note_no())

        self.date_edit = QDateEdit()
        self.date_edit.setDate(QDate.currentDate())
        self.date_edit.setCalendarPopup(True)

        self.tax_invoice_date = QDateEdit()
        self.tax_invoice_date.setDate(QDate.currentDate())
        self.tax_invoice_date.setCalendarPopup(True)

        self.invoice_input = QLineEdit()
        self.invoice_amount_input = QLineEdit("0.00")
        self.invoice_amount_input.textChanged.connect(self.update_amounts)

        self.customer_id = QLineEdit()
        self.customer_name = QLineEdit()
        self.customer_address = QTextEdit()
        self.customer_address.setFixedHeight(60)

        form_layout.addRow("Credit Note No:", self.credit_note_no)
        form_layout.addRow("Date:", self.date_edit)
        form_layout.addRow("Reference Invoice:", self.invoice_input)
        form_layout.addRow("Invoice Amount (Input):", self.invoice_amount_input)
        form_layout.addRow("Invoice Date:", self.tax_invoice_date)
        form_layout.addRow("Customer ID:", self.customer_id)
        form_layout.addRow("Customer Name:", self.customer_name)
        form_layout.addRow("Customer Address:", self.customer_address)

        self.table = QTableWidget(12, 4)
        self.table.setHorizontalHeaderLabels(["Item", "Quantity", "Unit Price", "Amount"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.cellChanged.connect(self.update_amounts)

        reason_layout = QFormLayout()
        self.reason = QLineEdit()

        self.total_credit = QLineEdit("0.00")
        self.total_credit.setReadOnly(True)

        self.different_amount = QLineEdit("0.00")
        self.different_amount.setReadOnly(True)

        self.vat_amount = QLineEdit("0.00")
        self.vat_amount.setReadOnly(True)

        self.total_with_vat = QLineEdit("0.00")
        self.total_with_vat.setReadOnly(True)

        self.invoice_amount_display = QLineEdit("0.00")
        self.invoice_amount_display.setReadOnly(True)

        self.thai_amount = QLineEdit()
        self.thai_amount.setReadOnly(True)

        reason_layout.addRow("Reason for Credit:", self.reason)
        reason_layout.addRow("Invoice Amount (Display):", self.invoice_amount_display)
        reason_layout.addRow("Total Credit Note:", self.total_credit)
        reason_layout.addRow("Different Amount:", self.different_amount)
        reason_layout.addRow("VAT 7%:", self.vat_amount)
        reason_layout.addRow("Total Amount (incl. VAT):", self.total_with_vat)
        reason_layout.addRow("Total Amount in Thai:", self.thai_amount)

        button_layout = QHBoxLayout()
        self.export_btn = QPushButton("Export Excel")
        self.export_btn.clicked.connect(self.export_to_excel)

        button_layout.addStretch()
        button_layout.addWidget(self.export_btn)

        layout.addLayout(form_layout)
        layout.addWidget(QLabel("Items:"))
        layout.addWidget(self.table)
        layout.addLayout(reason_layout)
        layout.addLayout(button_layout)

        self.setLayout(layout)
        self.update_amounts()

    def update_amounts(self):
        self.table.blockSignals(True)
        total = 0.0

        for row in range(self.table.rowCount()):
            item_cell = self.table.item(row, 0)
            item_text = item_cell.text().strip() if item_cell else ""

            qty_item = self.table.item(row, 1)
            price_item = self.table.item(row, 2)

            try:
                qty = float(qty_item.text()) if qty_item and qty_item.text().strip() else 0.0
                price = float(price_item.text()) if price_item and price_item.text().strip() else 0.0
            except ValueError:
                qty = 0.0
                price = 0.0

            if not item_text:
                amount_item = QTableWidgetItem("")
            else:
                amount = qty * price
                amount_item = QTableWidgetItem(f"{amount:.2f}")
                total += amount

            amount_item.setFlags(Qt.ItemIsEnabled)
            amount_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.table.setItem(row, 3, amount_item)

        self.total_credit.setText(to_arabic_digits(f"{total:.2f}"))

        try:
            invoice_amt = float(self.invoice_amount_input.text())
        except ValueError:
            invoice_amt = 0.0

        self.invoice_amount_display.setText(to_arabic_digits(f"{invoice_amt:.2f}"))
        diff = invoice_amt - total
        vat = diff * 0.07 if diff > 0 else 0.0
        total_vat = diff + vat if diff > 0 else diff

        self.different_amount.setText(to_arabic_digits(f"{diff:.2f}"))
        self.vat_amount.setText(to_arabic_digits(f"{vat:.2f}"))
        self.total_with_vat.setText(to_arabic_digits(f"{total_vat:.2f}"))

        try:
            thai_text = bahttext(float(f"{total_vat:.2f}"))
        except Exception:
            thai_text = ""
        self.thai_amount.setText(thai_text)

        self.table.blockSignals(False)

    def export_to_excel(self):
        credit_note_no = self.credit_note_no.text()
        file_path = os.path.join(self.output_dir, f"{credit_note_no}.xlsx")

        template_path = os.path.join(self.template_dir, "credit_note_template.xlsx")
        if not os.path.exists(template_path):
            QMessageBox.critical(self, "Error", f"Template not found at {template_path}")
            return

        accounting_format = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'

        try:
            wb = load_workbook(template_path)
            for sheet_name in ["ต้นฉบับ", "สำเนา"]:
                if sheet_name not in wb.sheetnames:
                    continue
                ws = wb[sheet_name]

                ws["H3"] = credit_note_no
                # Convert to Arabic digits string format dd/mm/yyyy + 543 year
                credit_note_date = self.date_edit.date().toPython()
                credit_note_date_str = f"{credit_note_date.day:02d}/{credit_note_date.month:02d}/{credit_note_date.year + 543}"
                ws["I9"] = to_arabic_digits(credit_note_date_str)

                ws["D12"] = self.invoice_input.text()
                ws["I26"] = self.invoice_amount_input.text()
                ws["D9"] = self.customer_id.text()
                ws["B10"] = self.customer_name.text()
                ws["B11"] = self.customer_address.toPlainText()
                ws["D32"] = self.reason.text()

                # Invoice date with same formatting + arabic digits
                invoice_date = self.tax_invoice_date.date().toPython()
                invoice_date_str = f"{invoice_date.day:02d}/{invoice_date.month:02d}/{invoice_date.year + 543}"
                ws["H12"] = to_arabic_digits(invoice_date_str)

                ws["I27"] = float(self.total_credit.text())
                ws["I27"].number_format = accounting_format

                ws["I28"] = float(self.different_amount.text())
                ws["I28"].number_format = accounting_format

                ws["I29"] = float(self.vat_amount.text())
                ws["I29"].number_format = accounting_format

                ws["I30"] = float(self.total_with_vat.text())
                ws["I30"].number_format = accounting_format

                ws["A30"] = self.thai_amount.text()

                start_row = 16
                for row in range(self.table.rowCount()):
                    item = self.table.item(row, 0)
                    qty = self.table.item(row, 1)
                    unit_price = self.table.item(row, 2)
                    amount = self.table.item(row, 3)

                    item_text = item.text() if item else ""
                    qty_text = qty.text() if qty else ""
                    unit_price_text = unit_price.text() if unit_price else ""
                    amount_text = amount.text() if amount else ""

                    if not any([item_text.strip(), qty_text.strip(), unit_price_text.strip(), amount_text.strip()]):
                        continue

                    ws[f"A{start_row}"] = float(qty_text or 0.0)
                    ws[f"A{start_row}"].number_format = accounting_format

                    ws[f"B{start_row}"] = item_text

                    ws[f"G{start_row}"] = float(unit_price_text or 0.0)
                    ws[f"G{start_row}"].number_format = accounting_format

                    ws[f"I{start_row}"] = float(amount_text or 0.0)
                    ws[f"I{start_row}"].number_format = accounting_format

                    start_row += 1

            wb.save(file_path)

            # Save JSON log
            now = datetime.now()
            log_entry = {
                "credit_note_no": credit_note_no,
                "date": now.strftime("%Y-%m-%d"),
                "time": now.strftime("%H:%M:%S"),
                "invoice_ref": self.invoice_input.text(),
                "customer_id": self.customer_id.text(),
                "customer_name": self.customer_name.text(),
                "total_with_vat": self.total_with_vat.text(),
            }
            self.save_credit_note_log(log_entry)

            QMessageBox.information(self, "Success", f"Exported to: {file_path}")

            # Prepare for next note
            self.credit_note_no.setText(self.generate_credit_note_no())

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Export failed:\n{str(e)}")

if __name__ == "__main__":
    app = QApplication([])
    window = CreditNoteWindow()
    app.setWindowIcon(QIcon(os.path.join(os.path.dirname(__file__), "credit_note.png")))
    window.show()
    app.exec()
